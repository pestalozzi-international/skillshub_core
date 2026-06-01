"""
Build Cohort 3 Student Upload spreadsheet for Frappe Data Import.

Cross-references Final Selection CSV (94 students) with Google Form
applications CSV (~398 entries) to produce an import-ready XLSX with
all SH Student fields including the new Application-tab fields.
"""

import csv
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent

FINAL_SEL_PATH = sys.argv[1] if len(sys.argv) > 1 else str(REPO_ROOT / "data" / "FINAL_SELECTION.csv")
APP_FORM_PATH = sys.argv[2] if len(sys.argv) > 2 else str(REPO_ROOT / "data" / "The_Edulutmation.csv")
OUT_PATH = sys.argv[3] if len(sys.argv) > 3 else str(REPO_ROOT / "outputs" / "Cohort3_Student_Upload.xlsx")


# ── Normalisation helpers ────────────────────────────────────────────────────


def norm(s):
	if not s:
		return ""
	s = str(s).strip().lower()
	s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
	return re.sub(r"\s+", " ", s)


def fname(full):
	p = str(full).strip().split()
	return p[0] if p else ""


def clean_phone(p):
	if not p:
		return ""
	p = re.sub(r"[^\d+]", "", str(p).split("/")[0])
	if p.startswith("0"):
		p = "+260" + p[1:]
	elif p.startswith("260") and not p.startswith("+"):
		p = "+" + p
	if p.startswith("+260260"):
		p = "+260" + p[7:]
	return p


GRADE_MAP = {
	"grade 6": "Grade 6",
	"grade 7": "Grade 7",
	"grade 8": "Grade 8",
	"grade 9": "Grade 9",
	"grade 10": "Grade 10",
	"grade 11": "Grade 11",
	"grade 12": "Grade 12",
	"grade 13": "Grade 12",
}


def clean_grade(g):
	return GRADE_MAP.get(norm(g), g.strip())


LEVEL_MAP = {
	"grade 1": "Grades 1 to 7",
	"grade 2": "Grades 1 to 7",
	"grade 3": "Grades 1 to 7",
	"grade 4": "Grades 1 to 7",
	"grade 5": "Grades 1 to 7",
	"grade 6": "Grades 1 to 7",
	"grade 7": "Grades 1 to 7",
	"grade 8": "Grades 8 to 9",
	"grade 9": "Grades 8 to 9",
	"grade 10": "Grades 10 to 12",
	"grade 11": "Grades 10 to 12",
	"grade 12": "Grades 10 to 12",
	"grade 13": "Grades 10 to 12",
}


def grade_to_level(g):
	return LEVEL_MAP.get(norm(g), "")


def clean_dob(d):
	if not d:
		return ""
	d = d.strip()
	if not re.match(r"^\d{4}-\d{2}-\d{2}$", d):
		return ""
	y = int(d[:4])
	if y < 1980 or y > 2015:
		return ""
	return d


def clean_gender(g):
	g = g.strip().lower()
	if g in ("male", "m"):
		return "Male"
	if g in ("female", "f"):
		return "Female"
	return g.capitalize()


def yes_no_to_check(v):
	return 1 if str(v).strip().lower() in ("yes", "y", "1") else 0


def clean_consent(v):
	v = str(v).strip().lower()
	if v in ("yes", "y", "1", "i agree"):
		return 1
	return 0


def clean_marital(v):
	v = str(v).strip().rstrip(".")
	m = {
		"single": "Single",
		"married": "Married",
		"divorced": "Divorced",
		"widowed": "Widowed",
		"other": "Other",
	}
	return m.get(v.lower().strip(), v.capitalize() if v else "")


REASON_MAP = {
	"financial difficulties": "Financial difficulties",
	"early parenthood": "Early parenthood",
	"family responsibilities": "Family responsibilities",
	"health issues": "Health issues",
	"lack of access to education": "Lack of access to education",
	"graduated": "Graduated",
	"other": "Other",
}


def clean_reason(v):
	return REASON_MAP.get(norm(v), v.strip() if v else "")


def parse_app_date(ts):
	if not ts:
		return ""
	try:
		dt = datetime.strptime(ts.split(" GMT")[0].strip(), "%Y/%m/%d %I:%M:%S %p")
		return dt.strftime("%Y-%m-%d")
	except ValueError:
		return ""


# ── Load Final Selection ─────────────────────────────────────────────────────

final_students = []
with open(FINAL_SEL_PATH, encoding="utf-8-sig") as f:
	rows = list(csv.reader(f))

for row in rows[2:]:

	def g(i, _row=row):
		return _row[i].strip() if i < len(_row) else ""

	if not g(0):
		continue
	final_students.append(
		{
			"full_name": g(0),
			"first_col": g(1),
			"gender": g(3),
			"area": g(4),
			"last_grade": g(5),
			"year_left": g(6),
			"employed": g(7),
			"is_parent": g(8),
			"trainee_num": g(9),
			"sel_pre_score": g(11),
			"numeracy": g(13),
			"literacy": g(14),
			"nf": norm(g(0)),
			"nfirst": norm(fname(g(0))),
		}
	)

print(f"Final selection loaded: {len(final_students)} students")

# ── Load Application Form ────────────────────────────────────────────────────

apps = []
with open(APP_FORM_PATH, encoding="utf-8-sig") as f:
	rows = list(csv.reader(f))

# Row 0-2 are headers (multi-row); data starts at row 3
for row in rows[3:]:

	def g(i, _row=row):
		return _row[i].strip() if i < len(_row) else ""

	full_name = g(2)
	if not full_name:
		continue
	apps.append(
		{
			"timestamp": g(0),
			"full_name": full_name,
			"dob": g(3),
			"age": g(4),
			"gender": g(5),
			"phone": g(6),
			"email": g(7),
			"emergency": g(8),
			"area": g(9),
			"last_grade": g(11),
			"last_school": g(12),
			"year_left": g(13),
			"reason_leaving": g(14),
			"marital_status": g(15),
			"employed": g(16),
			"employment_type": g(17),
			"is_parent": g(18),
			"num_children": g(19),
			"special_talents": g(20),
			"commitment_plan": g(21),
			"why_join": g(22),
			"what_achieve": g(23),
			"committed_duration": g(24),
			"challenges": g(25),
			"contact_consent": g(27),
			"media_consent": g(28),
			"declaration_consent": g(29),
			"nf": norm(full_name),
			"nfirst": norm(fname(full_name)),
		}
	)

print(f"Application forms loaded: {len(apps)} entries")


# ── Match function ───────────────────────────────────────────────────────────


def find_app(fs):
	for a in apps:
		if a["nf"] == fs["nf"]:
			return a, "exact"
	an = norm(fs["area"])
	m = [a for a in apps if a["nfirst"] == fs["nfirst"] and norm(a["area"]) == an]
	if len(m) == 1:
		return m[0], "first+area"
	if len(m) > 1:
		return m[-1], "first+area(last)"
	gn = norm(fs["last_grade"])
	m = [a for a in apps if a["nfirst"] == fs["nfirst"] and norm(a["last_grade"]) == gn]
	if len(m) == 1:
		return m[0], "first+grade"
	m = [a for a in apps if a["nfirst"] == fs["nfirst"]]
	if m:
		return m[-1], "first_only"
	return None, "no_match"


# ── Build rows ───────────────────────────────────────────────────────────────

rows_out = []
for fs in final_students:
	app, match_type = find_app(fs)

	parts = fs["full_name"].strip().split()
	fn = parts[0].capitalize() if parts else ""
	ln = " ".join(p.capitalize() for p in parts[1:]) if len(parts) > 1 else ""

	if app:
		phone = clean_phone(app["phone"])
		email = app["email"].strip()
		dob = clean_dob(app["dob"])
		area = (app["area"] or fs["area"]).strip()
		school = app["last_school"].strip()
		grade = clean_grade(app["last_grade"] or fs["last_grade"])
		level = grade_to_level(app["last_grade"] or fs["last_grade"])
		gender = clean_gender(app["gender"] or fs["gender"])
		emergency = app["emergency"].strip()
		year_left = (app["year_left"] or fs["year_left"]).strip()
		marital = clean_marital(app["marital_status"])
		employed_check = yes_no_to_check(app["employed"])
		emp_type = app["employment_type"].strip() if employed_check else ""
		is_parent_check = yes_no_to_check(app["is_parent"])
		num_children = app["num_children"].strip() if is_parent_check else ""
		reason = clean_reason(app["reason_leaving"])
		talents = app["special_talents"].strip()
		why_join = app["why_join"].strip()
		career_goals = app["what_achieve"].strip()
		challenges = app["challenges"].strip()
		media_consent = clean_consent(app["media_consent"])
		contact_consent = clean_consent(app["contact_consent"])
		app_date = parse_app_date(app["timestamp"])
		community = ""
	else:
		phone = email = dob = school = emergency = ""
		area = fs["area"].strip()
		grade = clean_grade(fs["last_grade"])
		level = grade_to_level(fs["last_grade"])
		gender = clean_gender(fs["gender"])
		year_left = fs["year_left"].strip()
		marital = ""
		employed_check = yes_no_to_check(fs["employed"])
		emp_type = ""
		is_parent_check = yes_no_to_check(fs["is_parent"])
		num_children = ""
		reason = talents = why_join = career_goals = challenges = community = ""
		media_consent = contact_consent = 0
		app_date = ""

	rows_out.append(
		{
			"first_name": fn,
			"last_name": ln,
			"student_name": fs["full_name"].strip(),
			"gender": gender,
			"date_of_birth": dob,
			"mobile": phone,
			"personal_email": email,
			"residential_area": area,
			"city": "Lusaka",
			"country": "Zambia",
			"last_school_attended": school,
			"last_year_of_schooling": grade,
			"highest_level_of_schooling": level,
			"students_occupation": "Yes" if fs["employed"].lower() in ("yes", "y") else "",
			"intake_cohort": "Cohort 3",
			"intake_year": "2026",
			"status": "Student",
			"naming_series": "SH.YY.####",
			# New fields
			"marital_status": marital,
			"is_parent": is_parent_check,
			"number_of_children": num_children,
			"emergency_contact": emergency,
			"currently_employed": employed_check,
			"employment_type": emp_type,
			"participation_challenges": challenges,
			"reason_for_leaving_school": reason,
			"special_talents": talents,
			"community_participation": community,
			"why_join_skillshub": why_join,
			"career_goals": career_goals,
			"how_skill_benefits_community": "",
			"how_skill_improves_livelihood": "",
			"preferred_course": "",
			"second_preference_course": "",
			"application_source": "Walk-in" if not app_date else "Online Portal",
			"application_date": app_date,
			"media_consent": media_consent,
			"contact_consent": contact_consent,
			# Reference-only columns
			"_trainee_num": fs["trainee_num"],
			"_numeracy": fs["numeracy"],
			"_literacy": fs["literacy"],
			"_match": match_type,
			"_is_parent": fs["is_parent"],
			"_sel_score": fs["sel_pre_score"],
		}
	)

matched = sum(1 for r in rows_out if r["_match"] != "no_match")
exact = sum(1 for r in rows_out if r["_match"] == "exact")
print(f"\nRows: {len(rows_out)} | Matched: {matched} (exact: {exact}) | No match: {len(rows_out) - matched}")


# ── Build XLSX ───────────────────────────────────────────────────────────────

wb = Workbook()

H_FILL = PatternFill("solid", start_color="1F4E79")
H_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
L_FILL = PatternFill("solid", start_color="BDD7EE")
D_FONT = Font(name="Arial", size=9)
A_FILL = PatternFill("solid", start_color="F2F2F2")
G_FILL = PatternFill("solid", start_color="E2EFDA")
Y_FILL = PatternFill("solid", start_color="FFF2CC")
R_FILL = PatternFill("solid", start_color="FCE4D6")
thin = Side(style="thin", color="D0D0D0")
BDR = Border(left=thin, right=thin, top=thin, bottom=thin)

MATCH_FILLS = {
	"exact": G_FILL,
	"first+area": G_FILL,
	"first+area(last)": Y_FILL,
	"first+grade": Y_FILL,
	"first_only": Y_FILL,
	"no_match": R_FILL,
}

# ── Sheet 1: Frappe Import ───────────────────────────────────────────────────

UPLOAD_COLS = [
	("naming_series", "Naming Series"),
	("first_name", "First Name"),
	("last_name", "Last Name"),
	("student_name", "Student Full Name"),
	("gender", "Gender"),
	("date_of_birth", "Date of Birth"),
	("mobile", "Mobile"),
	("personal_email", "Email"),
	("residential_area", "Residential Area"),
	("city", "City"),
	("country", "Country"),
	("last_school_attended", "Last School"),
	("last_year_of_schooling", "Last Year of Schooling"),
	("highest_level_of_schooling", "Highest Level"),
	("students_occupation", "Occupation at Selection"),
	("intake_cohort", "Intake Cohort"),
	("intake_year", "Intake Year"),
	("status", "Status"),
	("marital_status", "Marital Status"),
	("is_parent", "Is Parent?"),
	("number_of_children", "# Children"),
	("emergency_contact", "Emergency Contact"),
	("currently_employed", "Employed?"),
	("employment_type", "Employment Type"),
	("participation_challenges", "Participation Challenges"),
	("reason_for_leaving_school", "Reason Left School"),
	("special_talents", "Special Talents"),
	("community_participation", "Community Activities"),
	("why_join_skillshub", "Why Join Programme"),
	("career_goals", "Career Goals"),
	("how_skill_benefits_community", "Skill Benefits Community"),
	("how_skill_improves_livelihood", "Skill Improves Livelihood"),
	("preferred_course", "1st Course Pref"),
	("second_preference_course", "2nd Course Pref"),
	("application_source", "Application Source"),
	("application_date", "Application Date"),
	("media_consent", "Media Consent"),
	("contact_consent", "Contact Consent"),
]

ws1 = wb.active
ws1.title = "Frappe Import"

for ci, (field, _) in enumerate(UPLOAD_COLS, 1):
	c = ws1.cell(row=1, column=ci, value=field)
	c.font = Font(name="Arial", bold=True, color="1F4E79", size=9)
	c.fill = L_FILL
	c.border = BDR
	c.alignment = Alignment(horizontal="center", vertical="center")

ws1.row_dimensions[1].height = 18

for ri, rec in enumerate(rows_out, 2):
	fill = A_FILL if ri % 2 == 0 else PatternFill()
	for ci, (field, _) in enumerate(UPLOAD_COLS, 1):
		val = rec.get(field, "")
		c = ws1.cell(row=ri, column=ci, value=val)
		c.font = D_FONT
		c.border = BDR
		c.alignment = Alignment(vertical="center", wrap_text=False)
		if fill.fill_type:
			c.fill = fill

col_widths = [
	14,
	14,
	18,
	26,
	8,
	13,
	16,
	26,
	16,
	10,
	8,
	28,
	14,
	16,
	16,
	12,
	8,
	10,
	12,
	8,
	8,
	28,
	8,
	12,
	30,
	22,
	20,
	20,
	30,
	30,
	20,
	20,
	14,
	14,
	14,
	13,
	10,
	10,
]
for i, w in enumerate(col_widths, 1):
	if i <= len(col_widths):
		ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A2"

# ── Sheet 2: Reference & Audit ───────────────────────────────────────────────

ws2 = wb.create_sheet("Reference & Audit")

REF_COLS = [
	("student_name", "Full Name"),
	("_trainee_num", "Trainee #"),
	("_match", "App Form Match"),
	("first_name", "First Name"),
	("last_name", "Last Name"),
	("gender", "Gender"),
	("date_of_birth", "DOB"),
	("mobile", "Phone"),
	("personal_email", "Email"),
	("residential_area", "Area"),
	("last_school_attended", "Last School"),
	("last_year_of_schooling", "Last Grade"),
	("marital_status", "Marital Status"),
	("emergency_contact", "Emergency Contact"),
	("is_parent", "Parent?"),
	("number_of_children", "# Children"),
	("currently_employed", "Employed?"),
	("employment_type", "Emp Type"),
	("reason_for_leaving_school", "Reason Left"),
	("why_join_skillshub", "Why Join"),
	("career_goals", "Career Goals"),
	("_sel_score", "Sel Score"),
	("_numeracy", "Numeracy %"),
	("_literacy", "Literacy %"),
	("intake_cohort", "Cohort"),
	("status", "Status"),
]

for ci, (_, label) in enumerate(REF_COLS, 1):
	c = ws2.cell(row=1, column=ci, value=label)
	c.font = H_FONT
	c.fill = H_FILL
	c.border = BDR
	c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[1].height = 24

for ri, rec in enumerate(rows_out, 2):
	base_fill = A_FILL if ri % 2 == 0 else PatternFill()
	for ci, (field, _) in enumerate(REF_COLS, 1):
		val = rec.get(field, "")
		c = ws2.cell(row=ri, column=ci, value=val)
		c.font = D_FONT
		c.border = BDR
		c.alignment = Alignment(vertical="center")
		if field == "_match":
			c.fill = MATCH_FILLS.get(val, PatternFill())
			c.font = Font(name="Arial", size=9, bold=True)
		elif base_fill.fill_type:
			c.fill = base_fill

ref_widths = [
	26,
	12,
	18,
	14,
	18,
	8,
	13,
	16,
	26,
	14,
	28,
	14,
	12,
	28,
	8,
	8,
	8,
	10,
	20,
	30,
	30,
	10,
	10,
	10,
	12,
	8,
]
for i, w in enumerate(ref_widths, 1):
	ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ── Sheet 3: Instructions ────────────────────────────────────────────────────

ws3 = wb.create_sheet("Instructions")
ws3.sheet_view.showGridLines = False
instructions = [
	("HOW TO USE THIS SPREADSHEET", True, 14, "1F4E79"),
	("", False, 10, "000000"),
	("SHEET 1 - 'Frappe Import'", True, 11, "1F4E79"),
	("This sheet is formatted for Frappe/ERPNext Data Import (Insert new records).", False, 10, "000000"),
	("Steps:", True, 10, "000000"),
	("  1. In PI Portal: Settings > Data Import > New Data Import", False, 10, "000000"),
	("  2. Select DocType = 'SH Student', Import Type = 'Insert New Records'", False, 10, "000000"),
	("  3. Upload this sheet (row 1 = field names, rows 2+ = data)", False, 10, "000000"),
	("  4. Map columns if prompted, then Start Import", False, 10, "000000"),
	("", False, 10, "000000"),
	("SHEET 2 - 'Reference & Audit'", True, 11, "1F4E79"),
	("Shows all students with match quality and enriched data for review.", False, 10, "000000"),
	("The 'App Form Match' column colour codes:", False, 10, "000000"),
	("  GREEN (exact / first+area) = confident match from application form", False, 10, "000000"),
	("  YELLOW (first+grade / first_only) = best guess - please verify", False, 10, "000000"),
	("  RED (no_match) = no application form data found", False, 10, "000000"),
	("", False, 10, "000000"),
	("DATA SOURCES", True, 11, "1F4E79"),
	(
		"Final Selection CSV: Master list of selected students (name, area, grade, scores)",
		False,
		10,
		"000000",
	),
	(
		"Cohort 3 App Form: Full application data (DOB, phone, email, marital, employment, etc.)",
		False,
		10,
		"000000",
	),
	("", False, 10, "000000"),
	("FIELDS REQUIRING MANUAL COMPLETION", True, 11, "C00000"),
	("  - preferred_course / second_preference_course: set from course assignment", False, 10, "000000"),
	(
		"  - how_skill_benefits_community / how_skill_improves_livelihood: not in app form",
		False,
		10,
		"000000",
	),
	("  - DOB for no_match students: look up individually", False, 10, "000000"),
	("  - naming_series: Frappe auto-generates (SH.YY.####)", False, 10, "000000"),
	("", False, 10, "000000"),
	("COHORT NAMING", True, 11, "1F4E79"),
	("This file uses 'Cohort 3' which maps to the renamed cohort structure:", False, 10, "000000"),
	(
		"  Pilot Cohort (was Cohort 1), Cohort 1 (was 2), Cohort 2 (was 3), Cohort 3 (was 4)",
		False,
		10,
		"000000",
	),
	("Ensure the cohort rename is complete in the portal before importing.", False, 10, "000000"),
	("", False, 10, "000000"),
	(
		f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Total students: {len(rows_out)}",
		False,
		9,
		"666666",
	),
]

for ri, (text, bold, size, color) in enumerate(instructions, 1):
	c = ws3.cell(row=ri, column=1, value=text)
	c.font = Font(name="Arial", bold=bold, size=size, color=color)
	ws3.row_dimensions[ri].height = 16
ws3.column_dimensions["A"].width = 80

wb.save(OUT_PATH)
print(f"\nSaved: {OUT_PATH}")

# ── Summary ──────────────────────────────────────────────────────────────────

print("\n-- Match Quality --")
for mt in ["exact", "first+area", "first+area(last)", "first+grade", "first_only", "no_match"]:
	n = sum(1 for r in rows_out if r["_match"] == mt)
	print(f"  {mt:25s}: {n}")

no_match = [r for r in rows_out if r["_match"] == "no_match"]
if no_match:
	print("\nStudents needing manual DOB/data review:")
	for r in no_match:
		print(f"  {r['student_name']:30s}  trainee: {r['_trainee_num']}")
